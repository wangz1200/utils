# -*- coding: utf-8 -*-

import os
from collections import OrderedDict
import sqlalchemy as sa
from sqlalchemy.dialects.postgresql import Insert as insert
import openpyxl as xl

from db.db import DB


__day1__ = (31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31,)
__day2__ = (31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31,)

DEMAND = 1
FIX = 2
FIXES = {3: "03", 6: "06", 12: "12", 24: "24", 36: "36", 60: "60", }

DEMAND_BALANCE = "demand_balance"
DEMAND_MONTH_AVG = "demand_month_avg"
DEMAND_SEASON_AVG = "demand_season_avg"
DEMAND_YEAR_AVG = "demand_year_avg"

FIX_BALANCE = "fix_balance"
FIX_MONTH_AVG = "fix_month_avg"
FIX_SEASON_AVG = "fix_season_avg"
FIX_YEAR_AVG = "fix_year_avg"

SUM_BALANCE = "sum_balance"
SUM_MONTH_AVG = "sum_month_avg"
SUM_SEASON_AVG = "sum_season_avg"
SUM_YEAR_AVG = "sum_year_avg"


class Select(object):

    def __init__(self, db):
        super(Select, self).__init__()

        self.db = db
        self.sql = None

    def fields(self, *fields):
        self.sql = sa.select(fields)
        return self

    def select_from(self, from_):
        self.sql = self.sql.select_from(from_)
        return self

    def where(self, clause):
        self.sql = self.sql.where(clause)
        return self

    def group_by(self, *fields):
        self.sql = self.sql.group_by(*fields)
        return self


class Insert(object):

    def __init__(self, db):
        super(Insert, self).__init__()

        self.db = db
        self.t = None
        self.sql = None

    def table(self, table):
        self.t = self.db.table(table) if isinstance(table, str) else table
        self.sql = insert(self.t)
        return self

    def do_update(self):
        set_ = {}
        for c in self.t.c:
            set_[c.name] = getattr(self.sql.excluded, c.name)

        self.sql = self.sql.on_conflict_do_update(
            index_elements=self.t.primary_key,
            set_=set_
        )
        return self

    def do_nothing(self):
        self.sql = self.sql.on_conflict_do_nothing(
            index_elements=self.t.primary_key
        )
        return self

    def exec(self, *args, **kw):
        return self.db.commit(self.sql, *args, **kw)


def days(date):
    if len(date) != 8 or isinstance(date, str) is not True:
        return -1, -1, -1

    year = int(date[0:4])
    month = int(date[4:6])
    day = int(date[6:])

    days_ = __day2__ if year % 4 == 0 else __day1__

    m = day
    s = 0
    y = 0

    cur = 1
    while cur < month:
        y += days_[cur]
        cur = cur + 1
    y += day

    cur = month - (month - 1) % 3
    while cur < month:
        s += days_[cur]
        cur = cur + 1
    s += day

    return m, s, y


def register_user_table(db, name="user"):
    return db.register_table(
        name,
        sa.Column("user", sa.VARCHAR(32)),
        sa.Column("password", sa.VARCHAR(256)),
        sa.Column("name", sa.VARCHAR(32)),
        sa.Column("dept", sa.VARCHAR(32)),
        sa.Column("state", sa.VARCHAR(32)),
        sa.PrimaryKeyConstraint("user"),
    )


def register_customer_table(db, name="customer"):
    return db.register_table(
        name,
        sa.Column("customer", sa.VARCHAR(32)),
        sa.Column("name", sa.VARCHAR(128)),
        sa.Column("type", sa.VARCHAR(32)),
        sa.Column("open_date", sa.DATE),
        sa.PrimaryKeyConstraint("customer"),
    )


def register_deposit_account_table(db, name="deposit_account"):
    return db.register_table(
        name,
        sa.Column("account", sa.VARCHAR(32)),
        sa.Column("customer", sa.VARCHAR(32), ),
        sa.Column("inst", sa.VARCHAR(32)),
        sa.Column("product", sa.VARCHAR(32)),
        sa.Column("open_date", sa.DATE),
        sa.PrimaryKeyConstraint("account"),
        sa.ForeignKeyConstraint(("customer",), ["customer.customer", ]),
    )


def register_deposit_data_table(db, name="deposit_data"):
    return db.register_table(
        name,
        sa.Column("account", sa.VARCHAR(32)),
        sa.Column("state", sa.VARCHAR(32)),
        sa.Column("balance", sa.NUMERIC(32, 2, asdecimal=False)),
        sa.Column("month_acc", sa.NUMERIC(32, 2, asdecimal=False)),
        sa.Column("season_acc", sa.NUMERIC(32, 2, asdecimal=False)),
        sa.Column("year_acc", sa.NUMERIC(32, 2, asdecimal=False)),
        sa.Column("date", sa.DATE),
        sa.ForeignKeyConstraint(("account",), ["deposit_account.account", ]),
    )


def register_deposit_owner_table(db, name="deposit_owner"):
    return db.register_table(
        name,
        sa.Column("customer", sa.VARCHAR(32)),
        sa.Column("user", sa.VARCHAR(32)),
        sa.PrimaryKeyConstraint("customer"),
        sa.ForeignKeyConstraint(("user",), ["user.user", ]),
        sa.ForeignKeyConstraint(("customer",), ["customer.customer", ]),
    )


class SelectDeposit(object):

    def __init__(self, db, date, scale=10000, precision=2):
        super(SelectDeposit, self).__init__()

        self.db = db

        self.user = db.table("user")
        self.customer = db.table("customer")
        self.account = db.table("deposit_account")
        self.owner = db.table("deposit_owner")

        data = db.table("deposit_data")
        self.data = sa.select([data, ]).where(data.c.date == date).alias("data")

        self.source = self.data.join(self.account, self.data.c.account == self.account.c.account). \
            join(self.customer, self.customer.c.customer == self.account.c.customer). \
            outerjoin(self.owner, self.owner.c.customer == self.customer.c.customer). \
            outerjoin(self.user, self.user.c.user == self.owner.c.user)

        m, s, y = days(date)
        self.balance = sa.func.round(sa.func.sum(self.data.c.balance) / scale, precision)
        self.month_avg = sa.func.round(sa.func.sum(self.data.c.month_acc) / scale / m, precision)
        self.season_avg = sa.func.round(sa.func.sum(self.data.c.season_acc) / scale / s, precision)
        self.year_avg = sa.func.round(sa.func.sum(self.data.c.year_acc) / scale / y, precision)

        self.sql = None

    def where(self, clause):
        self.sql = self.sql.where(clause)
        return self

    def product(self, type_, *args):
        if type_ == DEMAND:
            self.where(sa.text(self.account.c.product.name + " ~ '%s'" % "^(?!113)"))

        elif type_ == FIX:
            where = "^113"
            if len(args) > 0:
                p = [FIXES[k] for k in args]
                if len(p) > 0:
                    where = where + "\\d{3}(%s)$" % "|".join(p)
            self.where(sa.text(self.account.c.product.name + " ~ '%s'" % where))

        return self

    def exec(self):
        return self.db.query(self.sql)


class SelectDepositByUser(SelectDeposit):

    def __init__(self, db, date):
        super(SelectDepositByUser, self).__init__(db, date)

        self.sql = sa.select([
            self.user.c.user.label("user"),
            self.user.c.name.label("name"),
            self.user.c.dept.label("dept"),
            self.user.c.state.label("state"),
            self.balance.label("balance"),
            self.month_avg.label("month_avg"),
            self.season_avg.label("season_avg"),
            self.year_avg.label("year_avg"),
        ]). \
            select_from(self.source). \
            group_by(self.user.c.user)


class SelectDepositByCustomer(SelectDeposit):

    def __init__(self, db, date):
        super(SelectDepositByCustomer, self).__init__(db, date)

        self.sql = sa.select([
            self.customer.c.customer.label("customer"),
            self.customer.c.name.label("name"),
            self.customer.c.type.label("type"),
            self.customer.c.open_date.label("open_date"),
            self.balance.label("balance"),
            self.month_avg.label("month_avg"),
            self.season_avg.label("season_avg"),
            self.year_avg.label("year_avg"),
            self.user.c.user.label("user"),
            self.user.c.name.label("user_name"),
            self.user.c.dept.label("dept"),
            self.user.c.state.label("state"),
        ]). \
            select_from(self.source). \
            group_by(self.customer.c.customer, self.user.c.user)


class MetaDeposit(object):

    def __init__(self, db, date, scale=10000, precision=2):
        super(MetaDeposit, self).__init__()

        self.db = db

        self.user = db.table("user")
        self.customer = db.table("customer")
        self.account = db.table("deposit_account")
        self.owner = db.table("deposit_owner")

        data = db.table("deposit_data")
        self.data = sa.select([data, ]).where(data.c.date == date).alias("data")

        self.source = self.data.join(self.account, self.data.c.account == self.account.c.account). \
            join(self.customer, self.customer.c.customer == self.account.c.customer). \
            outerjoin(self.owner, self.owner.c.customer == self.customer.c.customer). \
            outerjoin(self.user, self.user.c.user == self.owner.c.user)

        m, s, y = days(date)
        self.balance = sa.func.round(sa.func.sum(self.data.c.balance) / scale, precision)
        self.month_avg = sa.func.round(sa.func.sum(self.data.c.month_acc) / scale / m, precision)
        self.season_avg = sa.func.round(sa.func.sum(self.data.c.season_acc) / scale / s, precision)
        self.year_avg = sa.func.round(sa.func.sum(self.data.c.year_acc) / scale / y, precision)

    def product(self, type_, *args):
        if type_ == DEMAND:
            return sa.text(self.account.c.product.name + " ~ '%s'" % "^(?!113)")

        elif type_ == FIX:
            where = "^113"
            if len(args) > 0:
                p = [FIXES[k] for k in args]
                if len(p) > 0:
                    where = where + "\\d{3}(%s)$" % "|".join(p)
            return sa.text(self.account.c.product.name + " ~ '%s'" % where)

    def select_by_user(self):
        return sa.select([
            self.user.c.user.label("user"),
            self.user.c.name.label("name"),
            self.user.c.dept.label("dept"),
            self.user.c.state.label("state"),
            self.balance.label("balance"),
            self.month_avg.label("month_avg"),
            self.season_avg.label("season_avg"),
            self.year_avg.label("year_avg"),
        ]). \
            select_from(self.source). \
            group_by(self.user.c.user)

    def select_by_customer(self):
        return sa.select([
            self.customer.c.customer.label("customer"),
            self.customer.c.name.label("name"),
            self.customer.c.type.label("type"),
            self.customer.c.open_date.label("open_date"),
            self.balance.label("balance"),
            self.month_avg.label("month_avg"),
            self.season_avg.label("season_avg"),
            self.year_avg.label("year_avg"),
            self.user.c.user.label("user"),
            self.user.c.name.label("user_name"),
            self.user.c.dept.label("dept"),
            self.user.c.state.label("state"),
        ]). \
            select_from(self.source). \
            group_by(self.customer.c.customer, self.user.c.user)

    def select_by_account(self):
        return sa.select([
            self.account.c.account.label("account"),
            self.account.c.inst.label("inst"),
            self.account.c.product.label("product"),
            self.account.c.open_date.label("open_date"),
            self.customer.c.customer.label("customer"),
            self.customer.c.name.label("name"),
            self.customer.c.type.label("type"),
            self.customer.c.open_date.label("customer_open_date"),
            self.balance.label("balance"),
            self.month_avg.label("month_avg"),
            self.season_avg.label("season_avg"),
            self.year_avg.label("year_avg"),
            self.user.c.user.label("user"),
            self.user.c.name.label("user_name"),
            self.user.c.dept.label("dept"),
            self.user.c.state.label("state"),
        ]). \
            select_from(self.source). \
            group_by(self.account.c.account, self.customer.c.customer, self.user.c.user)

    def select_by_inst(self):
        return sa.select([
            self.account.c.inst.label("inst"),
            self.balance.label("balance"),
            self.month_avg.label("month_avg"),
            self.season_avg.label("season_avg"),
            self.year_avg.label("year_avg"),
        ]).\
            select_from(self.source).\
            group_by(self.account.c.inst)


class ExportDeposit(object):

    def __init__(self, meta):
        super(ExportDeposit, self).__init__()

        self.meta = meta
        self.sql = None


class ExportUserDeposit(object):

    def __init__(self, meta):
        super(ExportUserDeposit, self).__init__()

        self.meta = meta
        self.sql = self.meta.select_by_user()

    @classmethod
    def _row(cls, row):
        od = OrderedDict()

        od["user"] = row["user"]
        od["name"] = row["name"]
        od["dept"] = row["dept"]
        od["state"] = row["state"]

        od[DEMAND_BALANCE] = 0.00
        od[DEMAND_MONTH_AVG] = 0.00
        od[DEMAND_SEASON_AVG] = 0.00
        od[DEMAND_YEAR_AVG] = 0.00

        od[FIX_BALANCE] = 0.00
        od[FIX_MONTH_AVG] = 0.00
        od[FIX_SEASON_AVG] = 0.00
        od[FIX_YEAR_AVG] = 0.00

        od[SUM_BALANCE] = 0.00
        od[SUM_MONTH_AVG] = 0.00
        od[SUM_SEASON_AVG] = 0.00
        od[SUM_YEAR_AVG] = 0.00

        return od

    def _combine(self, demand, fix):
        res = {}

        for row in demand:
            user = row["user"]
            if user not in res:
                res[user] = self._row(row)
            item = res[user]
            item[DEMAND_BALANCE] = float(row["balance"])
            item[DEMAND_MONTH_AVG] = float(row["month_avg"])
            item[DEMAND_SEASON_AVG] = float(row["season_avg"])
            item[DEMAND_YEAR_AVG] = float(row["year_avg"])

        for row in fix:
            user = row["user"]
            if user not in res:
                res[user] = self._row(row)
            item = res[user]
            item[FIX_BALANCE] = float(row["balance"])
            item[FIX_MONTH_AVG] = float(row["month_avg"])
            item[FIX_SEASON_AVG] = float(row["season_avg"])
            item[FIX_YEAR_AVG] = float(row["year_avg"])

        for v in res.values():
            v[SUM_BALANCE] = v[DEMAND_BALANCE] + v[FIX_BALANCE]
            v[SUM_MONTH_AVG] = v[DEMAND_MONTH_AVG] + v[FIX_MONTH_AVG]
            v[SUM_SEASON_AVG] = v[DEMAND_SEASON_AVG] + v[FIX_SEASON_AVG]
            v[SUM_YEAR_AVG] = v[DEMAND_YEAR_AVG] + v[FIX_YEAR_AVG]

        return res

    def result(self):
        sql = self.sql.where(self.meta.product(DEMAND))
        demand = self.meta.db.query(sql)

        sql = self.sql.where(self.meta.product(FIX))
        fix = self.meta.db.query(sql)

        return self._combine(demand, fix)

    def save_to_excel(self, file, sheet=None, template=None):
        template = template or"./template/deposit/user.xlsx"

        wb = xl.load_workbook(template)
        ws = wb.active
        if sheet is not None:
            ws.title = sheet

        for row in self.result().values():
            ws.append(list(row.values()))

        wb.save(file)


class ExportCustomerDeposit(object):

    def __init__(self, meta):
        super(ExportCustomerDeposit, self).__init__()

        self.meta = meta
        self.sql = self.meta.select_by_customer()

    @classmethod
    def _row(cls, row):
        od = OrderedDict()

        od["customer"] = row["customer"]
        od["name"] = row["name"]
        od["type"] = row["type"]
        od["open_date"] = row["open_date"]

        od[DEMAND_BALANCE] = 0.00
        od[DEMAND_MONTH_AVG] = 0.00
        od[DEMAND_SEASON_AVG] = 0.00
        od[DEMAND_YEAR_AVG] = 0.00

        od[FIX_BALANCE] = 0.00
        od[FIX_MONTH_AVG] = 0.00
        od[FIX_SEASON_AVG] = 0.00
        od[FIX_YEAR_AVG] = 0.00

        od[SUM_BALANCE] = 0.00
        od[SUM_MONTH_AVG] = 0.00
        od[SUM_SEASON_AVG] = 0.00
        od[SUM_YEAR_AVG] = 0.00

        od["user"] = row["user"]
        od["user_name"] = row["user_name"]
        od["dept"] = row["dept"]
        od["state"] = row["state"]

        return od

    def _combine(self, demand, fix):
        res = {}

        for row in demand:
            customer = row["customer"]
            if customer not in res:
                res[customer] = self._row(row)
            item = res[customer]
            item[DEMAND_BALANCE] = float(row["balance"])
            item[DEMAND_MONTH_AVG] = float(row["month_avg"])
            item[DEMAND_SEASON_AVG] = float(row["season_avg"])
            item[DEMAND_YEAR_AVG] = float(row["year_avg"])

        for row in fix:
            customer = row["customer"]
            if customer not in res:
                res[customer] = self._row(row)
            item = res[customer]
            item[FIX_BALANCE] = float(row["balance"])
            item[FIX_MONTH_AVG] = float(row["month_avg"])
            item[FIX_SEASON_AVG] = float(row["season_avg"])
            item[FIX_YEAR_AVG] = float(row["year_avg"])

        for v in res.values():
            v[SUM_BALANCE] = v[DEMAND_BALANCE] + v[FIX_BALANCE]
            v[SUM_MONTH_AVG] = v[DEMAND_MONTH_AVG] + v[FIX_MONTH_AVG]
            v[SUM_SEASON_AVG] = v[DEMAND_SEASON_AVG] + v[FIX_SEASON_AVG]
            v[SUM_YEAR_AVG] = v[DEMAND_YEAR_AVG] + v[FIX_YEAR_AVG]

        return res

    def result(self):
        sql = self.sql.where(self.meta.product(DEMAND))
        demand = self.meta.db.query(sql)

        sql = self.sql.where(self.meta.product(FIX))
        fix = self.meta.db.query(sql)

        return self._combine(demand, fix)

    def save_to_excel(self, file, sheet=None, template=None):
        template = template or"./template/deposit/customer.xlsx"

        wb = xl.load_workbook(template)
        ws = wb.active
        if sheet is not None:
            ws.title = sheet

        for row in self.result().values():
            ws.append(list(row.values()))

        wb.save(file)


class ExportAccountDeposit(object):

    def __init__(self, meta):
        super(ExportAccountDeposit, self).__init__()

        self.meta = meta
        self.sql = self.meta.select_by_account()

    def result(self):
        res = {}

        for row in self.meta.db.query(self.sql):
            account = row["account"]

            od = OrderedDict()

            od["account"] = row["account"]
            od["inst"] = row["inst"]
            od["product"] = row["product"]
            od["open_date"] = row["open_date"]

            od["customer"] = row["customer"]
            od["name"] = row["name"]
            od["type"] = row["type"]
            od["customer_open_date"] = row["customer_open_date"]

            od["balance"] = float(row["balance"])
            od["month_avg"] = float(row["month_avg"])
            od["season_avg"] = float(row["season_avg"])
            od["year_avg"] = float(row["year_avg"])

            od["user"] = row["user"]
            od["user_name"] = row["user_name"]
            od["dept"] = row["dept"]
            od["state"] = row["state"]

            res[account] = od

        return res

    def save_to_excel(self, file, sheet=None, template=None):
        template = template or"./template/deposit/account.xlsx"

        wb = xl.load_workbook(template)
        ws = wb.active
        if sheet is not None:
            ws.title = sheet

        for row in self.result().values():
            ws.append(list(row.values()))

        wb.save(file)


class ExportInstDeposit(object):

    def __init__(self, meta):
        super(ExportInstDeposit, self).__init__()

        self.meta = meta
        self.sql = self.meta.select_by_inst()

    @classmethod
    def _row(cls, row):
        od = OrderedDict()

        od["inst"] = row["inst"]
        od["name"] = ""
        od["dept"] = ""
        od["state"] = ""

        od[DEMAND_BALANCE] = 0.00
        od[DEMAND_MONTH_AVG] = 0.00
        od[DEMAND_SEASON_AVG] = 0.00
        od[DEMAND_YEAR_AVG] = 0.00

        od[FIX_BALANCE] = 0.00
        od[FIX_MONTH_AVG] = 0.00
        od[FIX_SEASON_AVG] = 0.00
        od[FIX_YEAR_AVG] = 0.00

        od[SUM_BALANCE] = 0.00
        od[SUM_MONTH_AVG] = 0.00
        od[SUM_SEASON_AVG] = 0.00
        od[SUM_YEAR_AVG] = 0.00

        return od

    def _combine(self, demand, fix):
        res = {}

        for row in demand:
            inst = row["inst"]
            if inst not in res:
                res[inst] = self._row(row)
            item = res[inst]
            item[DEMAND_BALANCE] = float(row["balance"])
            item[DEMAND_MONTH_AVG] = float(row["month_avg"])
            item[DEMAND_SEASON_AVG] = float(row["season_avg"])
            item[DEMAND_YEAR_AVG] = float(row["year_avg"])

        for row in fix:
            inst = row["inst"]
            if inst not in res:
                res[inst] = self._row(row)
            item = res[inst]
            item[FIX_BALANCE] = float(row["balance"])
            item[FIX_MONTH_AVG] = float(row["month_avg"])
            item[FIX_SEASON_AVG] = float(row["season_avg"])
            item[FIX_YEAR_AVG] = float(row["year_avg"])

        for v in res.values():
            v[SUM_BALANCE] = v[DEMAND_BALANCE] + v[FIX_BALANCE]
            v[SUM_MONTH_AVG] = v[DEMAND_MONTH_AVG] + v[FIX_MONTH_AVG]
            v[SUM_SEASON_AVG] = v[DEMAND_SEASON_AVG] + v[FIX_SEASON_AVG]
            v[SUM_YEAR_AVG] = v[DEMAND_YEAR_AVG] + v[FIX_YEAR_AVG]

        return res

    def result(self):
        sql = self.sql.where(self.meta.product(DEMAND))
        demand = self.meta.db.query(sql)

        sql = self.sql.where(self.meta.product(FIX))
        fix = self.meta.db.query(sql)

        return self._combine(demand, fix)

    def save_to_excel(self, file, sheet=None, template=None):
        template = template or"./template/deposit/inst.xlsx"

        wb = xl.load_workbook(template)
        ws = wb.active
        if sheet is not None:
            ws.title = sheet

        for row in self.result().values():
            ws.append(list(row.values()))

        wb.save(file)


class ImportUser(object):

    def __init__(self, db):
        super(ImportUser, self).__init__()

        self.sql = Insert(db).table("user")

    def from_excel(self, file, sheet=None):
        wb = xl.load_workbook(file)
        ws = wb.active if sheet is None else wb[sheet]
        content = []
        for row in ws.iter_rows(min_row=2):
            content.append(
                {
                    "user": row[0].value,
                    "password": row[1].value,
                    "name": row[2].value,
                    "dept": row[3].value,
                    "state": row[4].value,
                }
            )
        wb.close()
        return self.sql.exec(*content)


class ImportCustomer(object):

    def __init__(self, db):
        super(ImportCustomer, self).__init__()

        self.sql = Insert(db).table("customer")

    def from_excel(self, file, sheet=None):
        wb = xl.load_workbook(file)
        ws = wb.active if sheet is None else wb[sheet]
        content = []
        for row in ws.iter_rows(min_row=2):
            content.append(
                {
                    "customer": row[4].value[-11:],
                    "name": row[5].value,
                    "type": row[9].value,
                    "open_date": row[10].value,
                }
            )
        wb.close()
        return self.sql.exec(*content)

    def from_txt(self, file, delimiter=","):
        content = []
        with open(file, encoding="utf8") as f:
            lines = f.readlines()
            if len(lines) <= 1:
                return
            for line in lines[1:]:
                row = line.split(delimiter)
                content.append({
                    "customer": row[4].replace(" ", "")[-11:],
                    "name": row[5].replace(" ", ""),
                    "type": row[9].replace(" ", ""),
                    "open_date": row[10].replace(" ", ""),
                })

        return self.sql.exec(*content)


class ImportDepositAccount(object):

    def __init__(self, db):
        super(ImportDepositAccount, self).__init__()

        self.sql = Insert(db).table("deposit_account")

    def from_excel(self, file, sheet=None):
        wb = xl.load_workbook(file)
        ws = wb.active if sheet is None else wb[sheet]
        content = []
        for row in ws.iter_rows(min_row=2):
            content.append(
                {
                    "account": row[3].value,
                    "customer": row[0].value[-11:],
                    "inst": row[1].value,
                    "product": row[7].value,
                    "open_date": row[9].value,
                }
            )
        wb.close()
        return self.sql.exec(*content)

    def from_txt(self, file, delimiter=","):
        content = []
        with open(file, encoding="utf8") as f:
            lines = f.readlines()
            for line in lines[1:]:
                row = line.split(delimiter)
                content.append({
                    "account": row[3].replace(" ", ""),
                    "customer": row[0].replace(" ", "")[-11:],
                    "inst": row[1].replace(" ", ""),
                    "product": row[7].replace(" ", ""),
                    "open_date": row[9].replace(" ", ""),
                })

        return self.sql.exec(*content)


class ImportDepositData(object):

    def __init__(self, db):
        super(ImportDepositData, self).__init__()

        self.sql = Insert(db).table("deposit_data")

    def from_excel(self, file, sheet=None):
        wb = xl.load_workbook(file)
        ws = wb.active if sheet is None else wb[sheet]
        content = []
        for row in ws.iter_rows(min_row=2):
            content.append(
                {
                    "account": row[3].value,
                    "state": row[11].value,
                    "balance": row[13].value,
                    "month_acc": row[14].value,
                    "season_acc": row[15].value,
                    "year_acc": row[16].value,
                    "date": row[17].value,
                }
            )
        wb.close()
        return self.sql.exec(*content)

    def from_txt(self, file, delimiter=","):
        content = []
        with open(file, encoding="utf8") as f:
            lines = f.readlines()
            for line in lines[1:]:
                row = line.split(delimiter)
                content.append({
                    "account": row[3].replace(" ", ""),
                    "state": row[11].replace(" ", ""),
                    "balance": row[13].replace(" ", ""),
                    "month_acc": row[14].replace(" ", ""),
                    "season_acc": row[15].replace(" ", ""),
                    "year_acc": row[16].replace(" ", ""),
                    "date": row[17].replace(" ", ""),
                })

        return self.sql.exec(*content)


class ImportDepositOwner(object):

    def __init__(self, db):
        super(ImportDepositOwner, self).__init__()

        self.sql = Insert(db).table("deposit_owner")

    def from_excel(self, file, sheet=None):
        wb = xl.load_workbook(file)
        ws = wb.active if sheet is None else wb[sheet]
        content = []
        for row in ws.iter_rows(min_row=2):
            content.append(
                {
                    "customer": row[0].value[-11:],
                    "user": row[1].value,
                }
            )
        wb.close()
        return self.sql.exec(*content)


def export_all_deposit(db, date, dir_):
    meta = MetaDeposit(db, date)

    ex = ExportDepositByUser(meta)
    ex.save_to_excel(os.path.join(dir_, "USER-" + date + ".xlsx"))

    ex = ExportDepositByCustomer(meta)
    ex.save_to_excel(os.path.join(dir_, "CUSTOMER-" + date + ".xlsx"))

    ex = ExportDepositByAccount(meta)
    ex.save_to_excel(os.path.join(dir_, "ACCOUNT-" + date + ".xlsx"))

    ex = ExportDepositByInst(meta)
    ex.sql = ex.sql.where(meta.user.c.user == None)
    ex.save_to_excel(os.path.join(dir_, "INST-" + date + ".xlsx"))


def import_all(db, dir_):
    files = os.listdir(dir_)
    for f in files:
        if f.startswith("CUS"):
            im = ImportCustomer(db)
            im.sql.do_nothing()
            im.from_txt(os.path.join(dir_, f))
        elif f.startswith("DEP"):
            im = ImportDepositAccount(db)
            im.sql.do_nothing()
            im.from_txt(os.path.join(dir_, f))

            im = ImportDepositData(db)
            im.from_txt(os.path.join(dir_, f))


def init(**kw):
    db = DB(**kw)

    register_user_table(db)
    register_customer_table(db)
    register_deposit_account_table(db)
    register_deposit_data_table(db)
    register_deposit_owner_table(db)

    return db


if __name__ == "__main__":

    db = init()

    date = "20190213"
    export_all_deposit(db, date, "D:/Desktop")

